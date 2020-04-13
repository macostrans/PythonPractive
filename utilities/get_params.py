import os
import openpyxl
import wget


class get_params:
    @staticmethod
    def get_params_from_gsheet(sheet_link, sheet_name):
        parameters_dict = {}
        parameters_list = []
        file = wget.download(sheet_link)
        book = openpyxl.load_workbook(file)
        sheet = book[sheet_name]
        for i in range(2, sheet.max_row + 1):  # to get rows
            parameters_dict = {}
            for j in range(1, sheet.max_column + 1):  # to get columns
                parameters_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            parameters_list.append(parameters_dict)
        os.remove(os.path.join(os.getcwd(), file))
        return parameters_list
